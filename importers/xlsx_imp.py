"""
Импортёры табличных данных (XLSX и CSV).

Каждый лист книги XLSX становится одним TableElement с подписью
«<имя файла> — <имя листа>». Перед таблицей вставляется заголовок 2-го
уровня, чтобы её можно было найти в содержании.

Для CSV — одна таблица, подпись = имя файла.
"""

import csv
import io
from typing import List, Optional

from importers import ImportResult
from elements import Heading, TableElement


def _strip_filename(filename: Optional[str]) -> str:
    name = filename or "data"
    if "/" in name:
        name = name.rsplit("/", 1)[-1]
    if "." in name:
        name = name.rsplit(".", 1)[0]
    return name


def _row_is_empty(row) -> bool:
    return all((c is None) or (str(c).strip() == "") for c in row)


def import_xlsx(data: bytes, filename: Optional[str] = None) -> ImportResult:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ImportResult(
            warnings=[
                "Для импорта XLSX установите openpyxl (pip install openpyxl)."
            ]
        )

    elements: List = []
    warnings: List[str] = []

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        return ImportResult(warnings=[f"XLSX: не удалось открыть файл — {e}"])

    base = _strip_filename(filename)

    for sheet in wb.worksheets:
        rows: List[List[str]] = []
        for raw_row in sheet.iter_rows(values_only=True):
            if _row_is_empty(raw_row):
                continue
            cells = ["" if c is None else str(c).strip() for c in raw_row]
            # Срезаем хвостовые пустые ячейки
            while cells and not cells[-1]:
                cells.pop()
            if cells:
                rows.append(cells)
        if not rows:
            continue
        caption = f"{base} — {sheet.title}" if len(wb.sheetnames) > 1 else base
        elements.append(TableElement(rows=rows, caption=caption, has_header=True))

    if not elements:
        warnings.append("XLSX: ни на одном листе не найдено данных.")

    return ImportResult(elements=elements, warnings=warnings)


def import_csv(data: bytes, filename: Optional[str] = None) -> ImportResult:
    elements: List = []
    warnings: List[str] = []

    # Пробуем декодировать с распространёнными кодировками
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return ImportResult(warnings=["CSV: не удалось декодировать файл."])

    # Автоопределение разделителя: ; , \t
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ","

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows: List[List[str]] = []
    for row in reader:
        cells = [c.strip() for c in row]
        while cells and not cells[-1]:
            cells.pop()
        if cells:
            rows.append(cells)

    if not rows:
        warnings.append("CSV: пустой файл.")
        return ImportResult(warnings=warnings)

    caption = _strip_filename(filename)
    elements.append(TableElement(rows=rows, caption=caption, has_header=True))

    return ImportResult(elements=elements, warnings=warnings)
